import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Write DataFrame to Exce
# Load the workbook
book = load_workbook('bibtexEntries.xlsx')

# Select the active sheet
sheet = book.active

# Iterate over the columns
for column in sheet.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Save the workbook
book.save('bibtexEntries.xlsx')
